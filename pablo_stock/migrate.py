import csv
import json
import os

import frappe
import openpyxl


# -----------------------------------------------------------------------------
# Helpers / Seeds
# -----------------------------------------------------------------------------

def ensure_warehouse_types():
	"""ERPNext usa Warehouse Type al crear warehouses por defecto al guardar Company."""
	required = ["Transit"]

	for wt in required:
		if not frappe.db.exists("Warehouse Type", wt):
			frappe.get_doc({
				"doctype": "Warehouse Type",
				"warehouse_type_name": wt,
			}).insert(ignore_permissions=True)


def ensure_uom(uom_name: str) -> str:
	"""Asegura un UOM existente para Items. Devuelve el name del UOM."""
	uom_name = (uom_name or "").strip()
	if not uom_name:
		return "Nos"

	if frappe.db.exists("UOM", uom_name):
		return uom_name

	doc = frappe.get_doc({
		"doctype": "UOM",
		"uom_name": uom_name,
		"enabled": 1,
	})
	doc.insert(ignore_permissions=True)
	return doc.name


def get_default_stock_uom() -> str:
	"""Intenta usar un UOM típico. Crea si falta."""
	for candidate in ("Nos", "Unit", "Units", "EA"):
		if frappe.db.exists("UOM", candidate):
			return candidate
	# si ninguno existe, crea "Nos"
	return ensure_uom("Nos")


def ensure_root_item_group() -> str:
	"""
	En ERPNext, el root suele ser 'All Item Groups'.
	Si por alguna razón no existe, lo crea.
	"""
	root = "All Item Groups"
	if frappe.db.exists("Item Group", root):
		return root

	doc = frappe.get_doc({
		"doctype": "Item Group",
		"item_group_name": root,
		"is_group": 1,
	})
	doc.insert(ignore_permissions=True)
	return doc.name


def load_items_sheet():
	"""Carga el Excel una sola vez y devuelve (sheet, headers_dict)."""
	file_path = frappe.get_app_path("pablo_stock", "nomenclators", "items.xlsx")
	workbook = openpyxl.load_workbook(file_path)
	sheet = workbook.active

	first_row = next(sheet.iter_rows(values_only=True))
	headers = [str(c).strip() if c is not None else "" for c in first_row]
	headers_dict = {h: i for i, h in enumerate(headers) if h}

	return sheet, headers_dict


def cell(row, headers_dict, key, default=None):
	"""Lee un valor de fila por nombre de header."""
	idx = headers_dict.get(key)
	if idx is None:
		return default
	return row[idx]


# -----------------------------------------------------------------------------
# Data loaders
# -----------------------------------------------------------------------------

def add_companys():
	with open(
		frappe.get_app_path("pablo_stock", "nomenclators", "company.csv"),
		newline="",
	) as data:
		companys = list(csv.reader(data, delimiter=","))

	for company in companys:
		name = (company[0] or "").strip().upper()
		abbr = (company[1] or "").strip().upper()
		currency = (company[2] or "").strip().upper()

		if not name or not abbr:
			continue

		# idempotente: por nombre y por abbr
		if frappe.db.exists("Company", name) or frappe.db.exists("Company", {"abbr": abbr}):
			continue

		doc = frappe.get_doc({
			"doctype": "Company",
			"company_name": name,
			"abbr": abbr,
			"default_currency": currency or "USD",
			"domain": "Pablo Stock",
			"country": "Uruguay",
		})
		doc.insert(ignore_permissions=True)


def add_brands():
	sheet, headers_dict = load_items_sheet()

	for row in sheet.iter_rows(values_only=True, min_row=2):
		brand = cell(row, headers_dict, "Meta: marca")
		if not brand:
			continue

		brand = str(brand).strip()
		if not brand or frappe.db.exists("Brand", brand):
			continue

		doc = frappe.get_doc({
			"doctype": "Brand",
			"brand": brand,
		})
		doc.insert(ignore_permissions=True)


def add_item_group():
	root = ensure_root_item_group()
	sheet, headers_dict = load_items_sheet()

	for row in sheet.iter_rows(values_only=True, min_row=2):
		item_group_name = cell(row, headers_dict, "Categorías")
		if not item_group_name:
			continue

		item_group_name = str(item_group_name).strip()
		if not item_group_name or frappe.db.exists("Item Group", item_group_name):
			continue

		item_group = frappe.get_doc({
			"doctype": "Item Group",
			"item_group_name": item_group_name,
			"parent_item_group": root,
			"is_group": 0,
		})
		item_group.insert(ignore_permissions=True)


def add_items():
	default_uom = get_default_stock_uom()

	sheet, headers_dict = load_items_sheet()

	for row in sheet.iter_rows(values_only=True, min_row=2):
		sku = cell(row, headers_dict, "SKU")
		if not sku:
			continue

		item_code = str(sku).strip()
		if not item_code or frappe.db.exists("Item", item_code):
			continue

		nombre = cell(row, headers_dict, "Nombre", "") or ""
		descripcion = cell(row, headers_dict, "Descripción", "") or ""
		categoria = cell(row, headers_dict, "Categorías", "") or ""
		modelo = cell(row, headers_dict, "Meta: modelo", "") or ""
		origen = cell(row, headers_dict, "Meta: origen", "") or ""
		aplica = cell(row, headers_dict, "Meta: aplica", "") or ""
		precio_normal = cell(row, headers_dict, "Precio normal", 0) or 0
		precio_rebajado = cell(row, headers_dict, "Precio rebajado", 0) or 0
		publicado = cell(row, headers_dict, "Publicado")

		# Brand
		marca = cell(row, headers_dict, "Meta: marca")
		brand_doc = ""
		if marca:
			marca = str(marca).strip()
			if marca and frappe.db.exists("Brand", marca):
				brand_doc = marca

		# Disabled
		disabled = 0
		if publicado is not None and str(publicado).strip().lower() in ("no", "false", "0"):
			disabled = 1

		# Item Group: si viene vacío o no existe, cae al root
		root_group = ensure_root_item_group()
		item_group = str(categoria).strip() if categoria else ""
		if not item_group or not frappe.db.exists("Item Group", item_group):
			item_group = root_group

		item = frappe.get_doc({
			"doctype": "Item",
			"item_code": item_code,
			"item_name": str(nombre).strip() or item_code,
			"description": str(descripcion).strip(),
			"item_group": item_group,

			# Requeridos típicos
			"stock_uom": default_uom,

			# Flags de negocio
			"disabled": disabled,
			"is_stock_item": 1,
			"is_sales_item": 1,
			"is_purchase_item": 1,

			# Campos custom tuyos
			"custom_origin": str(origen).strip(),
			"custom_item_model": str(modelo).strip(),
			"custom_applies": str(aplica).strip(),
			"custom_discounted_pr": precio_rebajado,

			# estándar (ojo: standard_rate no siempre se usa en ERPNext stock)
			"standard_rate": precio_normal,

			"brand": brand_doc,
		})
		item.insert(ignore_permissions=True)


# -----------------------------------------------------------------------------
# Roles / Permissions (tu lógica, solo ordenada)
# -----------------------------------------------------------------------------

def grant_permissions(role, items):
	for d in items:
		doc = frappe.new_doc("DocPerm")
		doc.role = role
		doc.parenttype = "DocType"
		doc.parentfield = "permissions"
		doc.update(d)


def copy_permissions(doctype):
	for d in frappe.get_all("DocPerm", {"parent": doctype}, "*"):
		if not frappe.db.exists(
			"Custom DocPerm",
			{"parent": d["parent"], "role": d["role"], "permlevel": d["permlevel"]},
		):
			doc = frappe.new_doc("Custom DocPerm")
			doc.update(d)
			doc.insert(ignore_permissions=True)


def grant_custom_permissions(role, items):
	for d in items:
		doc = frappe.new_doc("Custom DocPerm")
		doc.role = role
		doc.update(d)
		doc.insert(ignore_permissions=True)
		copy_permissions(d["parent"])


def has(role, items):
	for i in items:
		if i.get("parenttype") == "Report":
			frappe.db.delete("Has Role", {"role": role, "parenttype": "Report", "parent": i["parent"]})
			doc = frappe.new_doc("Has Role")
			doc.role = role
			doc.parentfield = "roles"
			doc.update(i)
			doc.insert(ignore_permissions=True)


def create_role(role):
	if frappe.db.exists("Role", role["name"]):
		doc = frappe.get_doc("Role", role["name"])
		doc.update(role)
		doc.save(ignore_permissions=True)
		frappe.db.delete("DocPerm", {"role": role["name"]})
		frappe.db.delete("Custom DocPerm", {"role": role["name"]})
	else:
		role["doctype"] = "Role"
		doc = frappe.get_doc(role)
		doc.insert(ignore_permissions=True)

	grant_permissions(role["name"], role.get("doc_perm", []))
	grant_custom_permissions(role["name"], role.get("custom_doc_perm", []))
	has(role["name"], role.get("has", []))


def restore_roles():
	folder = frappe.get_app_path("pablo_stock", "roles")
	for file_name in os.listdir(folder):
		if file_name.endswith(".json"):
			full_path = os.path.join(folder, file_name)
			with open(full_path) as f:
				create_role(json.load(f))


# -----------------------------------------------------------------------------
# Domain / Workspace
# -----------------------------------------------------------------------------

def add_domain():
	if not frappe.db.exists("Domain", {"domain": "Pablo Stock"}):
		domain_doc = frappe.new_doc("Domain")
		domain_doc.domain = "Pablo Stock"
		domain_doc.insert(ignore_permissions=True)

	if not frappe.db.exists("Has Domain", {"domain": "Pablo Stock"}):
		has_domain = frappe.new_doc("Has Domain")
		has_domain.domain = "Pablo Stock"
		has_domain.parent = "Domain Settings"
		has_domain.parenttype = "Domain Settings"
		has_domain.insert(ignore_permissions=True)


def disable_workspaces():
	modules_to_show = ["Pablo Stock", "Stock"]

	# Ocultar los que no son del módulo
	to_hide = frappe.get_all(
		"Workspace",
		filters={"module": ["not in", modules_to_show], "is_hidden": 0},
		pluck="name",
	)
	for name in to_hide:
		frappe.db.set_value("Workspace", name, "is_hidden", 1)

	# Mostrar los del módulo
	to_show = frappe.get_all(
		"Workspace",
		filters={"module": ["in", modules_to_show], "is_hidden": 1},
		pluck="name",
	)
	for name in to_show:
		frappe.db.set_value("Workspace", name, "is_hidden", 0)

	# Asegurar Users visible
	users_ws = frappe.get_all("Workspace", filters={"title": "Users", "is_hidden": 1}, pluck="name")
	for name in users_ws:
		frappe.db.set_value("Workspace", name, "is_hidden", 0)


# -----------------------------------------------------------------------------
# Entry point (hook)
# -----------------------------------------------------------------------------

def after_migrate():
	# Lo mínimo primero (para que Company no explote)
	ensure_warehouse_types()
	ensure_root_item_group()
	get_default_stock_uom()

	add_domain()
	restore_roles()

	add_companys()
	add_brands()
	add_item_group()
	add_items()

	disable_workspaces()
	frappe.db.commit()
